import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import io
import urllib.parse
import shutil
import os

EXCEL_FILE  = "Cust Details.xlsx"
BACKUP_FILE = "Cust Details_BACKUP.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# TRANSLATIONS  (English / Tamil)
# ══════════════════════════════════════════════════════════════════════════════
T = {
    "en": {
        "app_title":        "💎 Jewellery Management System",
        "menu":             "Menu",
        "select_page":      "Select Page",
        "language":         "🌐 Language",
        "pages":            ["Dashboard","Orders","Payments","Edit / Delete Order",
                             "Customer Ledger","Statement","Alerts","Settings","Help"],
        "login_title":      "🔐 Login",
        "login_user":       "Username",
        "login_pass":       "Password",
        "login_btn":        "Login",
        "login_err":        "❌ Wrong username or password.",
        "logout":           "🚪 Logout",
        "dashboard":        "Dashboard",
        "total_orders":     "Total Orders",
        "total_collection": "Total Collection",
        "total_balance":    "Total Balance Due",
        "ready_orders":     "Ready Orders",
        "partially_paid":   "Partially Paid",
        "completed":        "Completed (Paid)",
        "delivered":        "Delivered",
        "pending_delivery": "Pending Delivery",
        "pay_status_chart": "📊 Payment Status",
        "del_status_chart": "🚚 Delivery Status",
        "pending_alert":    "⚠️ Pending / Ready for Delivery",
        "no_pending":       "🎉 No pending deliveries!",
        "new_order":        "New Order",
        "cust_name":        "Customer Name",
        "phone":            "Phone Number",
        "address":          "Address",
        "item_name":        "Item Name",
        "item_desc":        "Item Description",
        "total_amt":        "Total Amount (₹)",
        "advance":          "Advance Paid (₹)",
        "order_date":       "Order Date",
        "delivery_date":    "Expected Delivery",
        "del_status":       "Delivery Status",
        "balance_preview":  "Balance after advance",
        "save_order":       "💾 Save Order",
        "jewel_details":    "💍 Jewellery Details",
        "metal_type":       "Metal Type",
        "purity":           "Purity / Karat",
        "weight":           "Weight (grams)",
        "making_charges":   "Making Charges (₹)",
        "stone_details":    "Stone / Diamond Details",
        "cost_price":       "Cost Price (₹)",
        "record_payment":   "Record Payment",
        "orders_balance":   "⏳ Orders with Balance Due",
        "no_balance":       "🎉 No pending balances!",
        "enter_payment":    "Enter Payment",
        "order_id":         "Order ID (e.g. O001)",
        "amount_now":       "Amount Being Paid Now (₹)",
        "payment_date":     "Payment Date",
        "payment_mode":     "Payment Mode",
        "txn_id":           "Transaction ID (optional)",
        "save_payment":     "💾 Save Payment",
        "edit_delete":      "Edit / Delete Order",
        "edit_info":        "Search an Order ID to edit its details or delete it.",
        "search_order":     "🔍 Enter Order ID",
        "editing":          "Editing Order",
        "save_changes":     "💾 Save Changes",
        "delete_order":     "🗑️ Delete Order",
        "delete_warn":      "⚠️ This will permanently delete this order and all payments!",
        "confirm_del":      "Type Order ID to confirm deletion",
        "cust_ledger":      "Customer Ledger",
        "ledger_info":      "Search any customer to see full order & payment history.",
        "search_cust":      "🔍 Search by Name / Phone / Order ID",
        "no_cust":          "No customer found.",
        "pay_history":      "Payment History",
        "no_pay_rec":       "No payment records for this order.",
        "download_receipt": "🖨️ Download Receipt",
        "wa_reminder":      "📲 Send WhatsApp Reminder",
        "wa_enter_num":     "Enter Customer WhatsApp Number",
        "wa_open":          "✅ Open WhatsApp & Send",
        "wa_caption":       "WhatsApp will open with message typed. Just press Send!",
        "wa_invalid":       "Enter a valid 10-digit number.",
        "fully_paid":       "✅ Fully paid!",
        "still_pending":    "⚠️ still pending from",
        "statement":        "Statement",
        "order_details":    "📋 Order Details",
        "payments_for":     "💳 Payments for",
        "no_pay_found":     "No payment records found.",
        "quick_summary":    "📊 Quick Summary",
        "total_in_system":  "Total Orders in System",
        "recent_orders":    "Recent Orders",
        "settings":         "Settings",
        "shop_name":        "Shop Name",
        "shop_phone":       "Shop Phone",
        "shop_address":     "Shop Address",
        "settings_info":    "To change shop details, edit the 'Settings' sheet in Excel.",
        "backup_restore":   "Backup & Restore",
        "backup_now":       "📥 Backup Excel Now",
        "backup_success":   "✅ Backup created!",
        "backup_download":  "⬇️ Download Backup File",
        "restore_title":    "♻️ Restore from Backup",
        "restore_upload":   "Upload a backup .xlsx file to restore",
        "restore_btn":      "✅ Restore Now",
        "restore_success":  "✅ Data restored successfully!",
        "restore_warn":     "⚠️ This will overwrite current data!",
        "help":             "Help",
    },
    "ta": {
        "app_title":        "💎 நகை மேலாண்மை அமைப்பு",
        "menu":             "பட்டி",
        "select_page":      "பக்கம் தேர்வு செய்யவும்",
        "language":         "🌐 மொழி",
        "pages":            ["டாஷ்போர்டு","ஆர்டர்கள்","கட்டணங்கள்","திருத்து / நீக்கு",
                             "வாடிக்கையாளர் ஏடு","அறிக்கை","விழிப்பூட்டல்கள்","அமைப்புகள்","உதவி"],
        "login_title":      "🔐 உள்நுழைவு",
        "login_user":       "பயனர் பெயர்",
        "login_pass":       "கடவுச்சொல்",
        "login_btn":        "உள்நுழை",
        "login_err":        "❌ தவறான பயனர் பெயர் அல்லது கடவுச்சொல்.",
        "logout":           "🚪 வெளியேறு",
        "dashboard":        "டாஷ்போர்டு",
        "total_orders":     "மொத்த ஆர்டர்கள்",
        "total_collection": "மொத்த வசூல்",
        "total_balance":    "மொத்த நிலுவை",
        "ready_orders":     "தயார் ஆர்டர்கள்",
        "partially_paid":   "பகுதியாக செலுத்தியது",
        "completed":        "நிறைவு (செலுத்தியது)",
        "delivered":        "டெலிவரி ஆனது",
        "pending_delivery": "நிலுவை டெலிவரி",
        "pay_status_chart": "📊 கட்டண நிலை",
        "del_status_chart": "🚚 டெலிவரி நிலை",
        "pending_alert":    "⚠️ நிலுவை / டெலிவரிக்கு தயார்",
        "no_pending":       "🎉 நிலுவை டெலிவரி இல்லை!",
        "new_order":        "புதிய ஆர்டர்",
        "cust_name":        "வாடிக்கையாளர் பெயர்",
        "phone":            "தொலைபேசி எண்",
        "address":          "முகவரி",
        "item_name":        "பொருள் பெயர்",
        "item_desc":        "பொருள் விவரம்",
        "total_amt":        "மொத்த தொகை (₹)",
        "advance":          "முன்பணம் (₹)",
        "order_date":       "ஆர்டர் தேதி",
        "delivery_date":    "டெலிவரி தேதி",
        "del_status":       "டெலிவரி நிலை",
        "balance_preview":  "முன்பணம் பிறகு நிலுவை",
        "save_order":       "💾 ஆர்டர் சேமி",
        "jewel_details":    "💍 நகை விவரங்கள்",
        "metal_type":       "உலோக வகை",
        "purity":           "தூய்மை / கேரட்",
        "weight":           "எடை (கிராம்)",
        "making_charges":   "செய்கூலி (₹)",
        "stone_details":    "கல் / வைரம் விவரங்கள்",
        "cost_price":       "கொள்முதல் விலை (₹)",
        "record_payment":   "கட்டணம் பதிவு செய்",
        "orders_balance":   "⏳ நிலுவை உள்ள ஆர்டர்கள்",
        "no_balance":       "🎉 நிலுவை கட்டணம் இல்லை!",
        "enter_payment":    "கட்டணம் உள்ளிடவும்",
        "order_id":         "ஆர்டர் ID (எ.கா. O001)",
        "amount_now":       "இப்போது செலுத்தும் தொகை (₹)",
        "payment_date":     "கட்டண தேதி",
        "payment_mode":     "கட்டண முறை",
        "txn_id":           "பரிவர்த்தனை ID (விரும்பினால்)",
        "save_payment":     "💾 கட்டணம் சேமி",
        "edit_delete":      "திருத்து / நீக்கு ஆர்டர்",
        "edit_info":        "ஆர்டர் திருத்த அல்லது நீக்க ID தேடவும்.",
        "search_order":     "🔍 ஆர்டர் ID உள்ளிடவும்",
        "editing":          "திருத்துகிறது ஆர்டர்",
        "save_changes":     "💾 மாற்றங்கள் சேமி",
        "delete_order":     "🗑️ ஆர்டர் நீக்கு",
        "delete_warn":      "⚠️ இது ஆர்டர் மற்றும் கட்டணங்களை நிரந்தரமாக நீக்கும்!",
        "confirm_del":      "உறுதிப்படுத்த ஆர்டர் ID தட்டச்சு செய்யவும்",
        "cust_ledger":      "வாடிக்கையாளர் ஏடு",
        "ledger_info":      "வாடிக்கையாளர் ஆர்டர் மற்றும் கட்டண வரலாற்றை காணவும்.",
        "search_cust":      "🔍 பெயர் / தொலைபேசி / ஆர்டர் ID தேடவும்",
        "no_cust":          "வாடிக்கையாளர் கிடைக்கவில்லை.",
        "pay_history":      "கட்டண வரலாறு",
        "no_pay_rec":       "கட்டண பதிவுகள் இல்லை.",
        "download_receipt": "🖨️ ரசீது பதிவிறக்கம்",
        "wa_reminder":      "📲 WhatsApp நினைவூட்டல்",
        "wa_enter_num":     "வாடிக்கையாளர் WhatsApp எண் உள்ளிடவும்",
        "wa_open":          "✅ WhatsApp திறந்து அனுப்பு",
        "wa_caption":       "WhatsApp செய்தி தயாராக இருக்கும். அனுப்பு அழுத்தவும்!",
        "wa_invalid":       "சரியான 10 இலக்க எண் உள்ளிடவும்.",
        "fully_paid":       "✅ முழுவதும் செலுத்தியது!",
        "still_pending":    "⚠️ நிலுவை உள்ளது",
        "statement":        "அறிக்கை",
        "order_details":    "📋 ஆர்டர் விவரங்கள்",
        "payments_for":     "💳 கட்டணங்கள்",
        "no_pay_found":     "கட்டண பதிவுகள் இல்லை.",
        "quick_summary":    "📊 சுருக்கம்",
        "total_in_system":  "மொத்த ஆர்டர்கள்",
        "recent_orders":    "சமீபத்திய ஆர்டர்கள்",
        "settings":         "அமைப்புகள்",
        "shop_name":        "கடை பெயர்",
        "shop_phone":       "கடை தொலைபேசி",
        "shop_address":     "கடை முகவரி",
        "settings_info":    "கடை விவரங்களை Excel 'Settings' தாளில் மாற்றவும்.",
        "backup_restore":   "காப்புப்பிரதி & மீட்டமை",
        "backup_now":       "📥 இப்போது காப்புப்பிரதி எடு",
        "backup_success":   "✅ காப்புப்பிரதி உருவாக்கப்பட்டது!",
        "backup_download":  "⬇️ காப்புப்பிரதி பதிவிறக்கம்",
        "restore_title":    "♻️ காப்புப்பிரதியிலிருந்து மீட்டமை",
        "restore_upload":   "மீட்டமைக்க .xlsx கோப்பை பதிவேற்றவும்",
        "restore_btn":      "✅ மீட்டமை",
        "restore_success":  "✅ தரவு வெற்றிகரமாக மீட்டமைக்கப்பட்டது!",
        "restore_warn":     "⚠️ இது தற்போதைய தரவை மேலெழுதும்!",
        "help":             "உதவி",
    }
}

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def safe_num(val):
    try:
        if val is None: return 0
        if str(val).startswith("="): return 0
        return float(val)
    except (ValueError, TypeError):
        return 0

def fmt_date(val):
    if val is None: return ""
    s = str(val)
    return s.split(" ")[0] if " " in s else s

def get_shop_info():
    try:
        wb  = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        ws  = wb["Settings"]
        row = list(ws.iter_rows(min_row=2, values_only=True))[0]
        wb.close()
        return {"name": str(row[0] or "Jewellery Shop"),
                "phone": str(row[1] or ""),
                "address": str(row[2] or "")}
    except Exception:
        return {"name": "Jewellery Shop", "phone": "", "address": ""}

def t(key):
    """Get translation for current language."""
    lang = st.session_state.get("lang", "en")
    return T[lang].get(key, T["en"].get(key, key))

def generate_pdf_receipt(order_data, payments):
    shop  = get_shop_info()
    o     = order_data
    total = safe_num(o[9])
    paid  = safe_num(o[10])
    bal   = total - paid
    pay_rows = ""
    for p in payments:
        pay_rows += f"<tr><td>{fmt_date(p[2])}</td><td>₹{safe_num(p[3]):,.0f}</td><td>{p[4] or ''}</td><td>{p[5] or ''}</td></tr>"
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body{{font-family:Arial,sans-serif;margin:40px;color:#222}}
  .header{{text-align:center;border-bottom:2px solid #c8a96e;padding-bottom:12px;margin-bottom:20px}}
  .header h1{{margin:0;color:#8B5E3C;font-size:26px}}
  .header p{{margin:4px 0;font-size:13px;color:#555}}
  .section{{margin-bottom:18px}}
  .section h3{{background:#f5efe6;padding:6px 10px;border-left:4px solid #c8a96e;margin:0 0 8px 0;font-size:14px;color:#8B5E3C}}
  table{{width:100%;border-collapse:collapse;font-size:13px}}
  td,th{{padding:7px 10px;border:1px solid #ddd}}
  th{{background:#f5efe6;color:#8B5E3C;text-align:left}}
  .label{{font-weight:bold;width:40%;background:#fafafa}}
  .total-row{{background:#fff8ee;font-weight:bold}}
  .balance{{font-size:18px;font-weight:bold;color:{'#d9534f' if bal>0 else '#28a745'};text-align:right;margin-top:10px}}
  .footer{{text-align:center;font-size:11px;color:#999;margin-top:30px;border-top:1px solid #eee;padding-top:10px}}
  @media print{{body{{margin:20px}}}}
</style></head><body>
<div class="header">
  <h1>💎 {shop['name']}</h1>
  <p>{shop['address']}</p><p>📞 {shop['phone']}</p>
  <p style="font-size:11px;color:#aaa">Receipt generated on {date.today()}</p>
</div>
<div class="section"><h3>ORDER DETAILS</h3><table>
  <tr><td class="label">Order ID</td><td>{o[0]}</td><td class="label">Customer ID</td><td>{o[1]}</td></tr>
  <tr><td class="label">Customer Name</td><td>{o[2]}</td><td class="label">Phone</td><td>{o[3]}</td></tr>
  <tr><td class="label">Address</td><td colspan="3">{o[4] or ''}</td></tr>
  <tr><td class="label">Item Name</td><td>{o[5]}</td><td class="label">Description</td><td>{o[6] or ''}</td></tr>
  <tr><td class="label">Order Date</td><td>{fmt_date(o[7])}</td><td class="label">Delivery Date</td><td>{fmt_date(o[8])}</td></tr>
  <tr><td class="label">Payment Status</td><td>{o[12] or ''}</td><td class="label">Delivery Status</td><td>{o[13] or ''}</td></tr>
</table></div>
<div class="section"><h3>PAYMENT HISTORY</h3><table>
  <tr><th>Date</th><th>Amount</th><th>Mode</th><th>Transaction ID</th></tr>
  {pay_rows if pay_rows else '<tr><td colspan="4" style="text-align:center">No payments recorded</td></tr>'}
</table></div>
<div class="section"><h3>SUMMARY</h3><table>
  <tr class="total-row"><td class="label">Total Amount</td><td>₹{total:,.0f}</td></tr>
  <tr><td class="label">Total Paid</td><td>₹{paid:,.0f}</td></tr>
  <tr class="total-row"><td class="label">Balance Due</td><td>₹{bal:,.0f}</td></tr>
</table>
<p class="balance">{'⚠️ Balance Due: ₹' + f'{bal:,.0f}' if bal>0 else '✅ FULLY PAID'}</p>
</div>
<div class="footer">Thank you for your business! — {shop['name']}</div>
</body></html>"""
    return html

def whatsapp_link(manual_phone, order_id, customer_name, item_name, balance, delivery_date, shop_name):
    clean = "".join(filter(str.isdigit, str(manual_phone)))
    if len(clean) == 10:
        clean = "91" + clean
    msg = (
        f"Dear {customer_name},\n\n"
        f"This is a reminder from *{shop_name}*.\n\n"
        f"📦 Order ID   : {order_id}\n"
        f"💍 Item       : {item_name}\n"
        f"📅 Delivery   : {fmt_date(delivery_date)}\n"
        f"💰 Balance Due: ₹{balance:,.0f}\n\n"
        f"Kindly clear the pending balance at the time of delivery.\n\n"
        f"Thank you! 🙏"
    )
    return f"https://wa.me/{clean}?text={urllib.parse.quote(msg)}"

def whatsapp_widget(order_id, customer_name, item_name, balance, delivery_date, shop_name, key_suffix=""):
    with st.expander(t("wa_reminder"), expanded=False):
        wa_num = st.text_input(t("wa_enter_num"), placeholder="e.g. 9876543210", key=f"wa_num_{key_suffix}")
        if wa_num:
            clean = "".join(filter(str.isdigit, wa_num))
            if len(clean) not in (10, 12):
                st.warning(t("wa_invalid"))
            else:
                wa_url = whatsapp_link(wa_num, order_id, customer_name, item_name, balance, delivery_date, shop_name)
                st.link_button(t("wa_open"), wa_url)
                st.caption(t("wa_caption"))

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════════════
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "lang" not in st.session_state:
    st.session_state["lang"] = "en"

# ══════════════════════════════════════════════════════════════════════════════
# LOGIN PAGE  —  Role-based access
# ══════════════════════════════════════════════════════════════════════════════
# Roles: admin = full access | manager = no delete | staff = orders+payments only
USERS = {
    "admin":   {"password": "jewel123",  "role": "admin"},
    "owner":   {"password": "shop456",   "role": "admin"},
    "manager": {"password": "mgr789",    "role": "manager"},
    "staff":   {"password": "staff111",  "role": "staff"},
}

# Pages allowed per role
ROLE_PAGES = {
    "admin":   ["Dashboard","Orders","Payments","Edit / Delete Order",
                "Customer Ledger","Statement","Alerts","Settings","Help"],
    "manager": ["Dashboard","Orders","Payments","Edit / Delete Order",
                "Customer Ledger","Statement","Alerts","Help"],
    "staff":   ["Orders","Payments","Customer Ledger","Help"],
}

st.set_page_config(page_title="Jewellery Management System", layout="wide")

if not st.session_state["logged_in"]:
    st.markdown("<br><br>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown("<h2 style='text-align:center;color:#8B5E3C'>💎 Jewellery Management System</h2>", unsafe_allow_html=True)
        st.markdown(f"### {t('login_title')}")
        uname = st.text_input(t("login_user"), key="login_uname")
        passw = st.text_input(t("login_pass"), type="password", key="login_pass")
        lang_choice = st.radio("🌐 Language / மொழி", ["English", "தமிழ்"], horizontal=True, key="login_lang")
        st.session_state["lang"] = "ta" if lang_choice == "தமிழ்" else "en"
        if st.button(t("login_btn"), type="primary", use_container_width=True):
            user = USERS.get(uname)
            if user and user["password"] == passw:
                st.session_state["logged_in"] = True
                st.session_state["username"]  = uname
                st.session_state["role"]      = user["role"]
                st.rerun()
            else:
                st.error(t("login_err"))
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR  (only shown after login)
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.title(t("app_title"))

    # Language toggle
    lang_choice = st.radio(t("language"), ["English", "தமிழ்"], horizontal=True,
                           index=0 if st.session_state["lang"]=="en" else 1)
    st.session_state["lang"] = "ta" if lang_choice == "தமிழ்" else "en"

    # Filter pages based on role
    role       = st.session_state.get("role","staff")
    allowed_en = ROLE_PAGES.get(role, ROLE_PAGES["staff"])
    all_en     = T["en"]["pages"] + ["Alerts"]   # master list
    # Add Alerts to translations if missing
    all_ta     = T["ta"]["pages"] + ["விழிப்பூட்டல்கள்"]
    lang       = st.session_state.get("lang","en")
    all_local  = all_ta if lang=="ta" else all_en
    # Build filtered list in local language
    filtered_local = []
    filtered_en    = []
    for i, en_pg in enumerate(all_en):
        if en_pg in allowed_en:
            filtered_en.append(en_pg)
            filtered_local.append(all_local[i] if i < len(all_local) else en_pg)

    page_key = st.selectbox(t("select_page"), range(len(filtered_local)),
                             format_func=lambda i: filtered_local[i])
    page     = filtered_en[page_key]

    st.divider()
    role_badge = {"admin":"👑 Admin","manager":"🔧 Manager","staff":"👤 Staff"}.get(role, role)
    st.caption(f"{role_badge} — {st.session_state.get('username','')}")
    if st.button(t("logout")):
        st.session_state["logged_in"] = False
        st.rerun()

st.title(t("app_title"))

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
page_en_name = page   # page is already English key from sidebar

if page_en_name == "Dashboard":
    st.header(t("dashboard"))
    try:
        wb     = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        orders = list(wb["Orders"].iter_rows(min_row=2, values_only=True))
        wb.close()
        total_orders = total_collection = total_balance = 0
        pay_counts = {"Paid":0,"Partially Paid":0,"Pending":0}
        del_counts  = {"Delivered":0,"Ready":0,"Pending":0}
        for r in orders:
            if not r[0]: continue
            total_orders     += 1
            total_collection += safe_num(r[9])
            total_balance    += max(safe_num(r[9])-safe_num(r[10]), 0)
            ps = str(r[12]).strip() if r[12] else "Pending"
            ds = str(r[13]).strip() if r[13] else "Pending"
            pay_counts[ps] = pay_counts.get(ps,0)+1
            del_counts[ds] = del_counts.get(ds,0)+1

        c1,c2,c3,c4 = st.columns(4)
        c1.metric(t("total_orders"),     total_orders)
        c2.metric(t("total_collection"), f"₹{total_collection:,.0f}")
        c3.metric(t("total_balance"),    f"₹{total_balance:,.0f}")
        c4.metric(t("ready_orders"),     del_counts.get("Ready",0))
        c5,c6,c7,c8 = st.columns(4)
        c5.metric(t("partially_paid"),   pay_counts.get("Partially Paid",0))
        c6.metric(t("completed"),        pay_counts.get("Paid",0))
        c7.metric(t("delivered"),        del_counts.get("Delivered",0))
        c8.metric(t("pending_delivery"), del_counts.get("Pending",0))
        st.divider()
        ch1, ch2 = st.columns(2)
        with ch1:
            st.subheader(t("pay_status_chart"))
            df_pc = pd.DataFrame({"Status":list(pay_counts.keys()),"Count":list(pay_counts.values())})
            st.bar_chart(df_pc[df_pc["Count"]>0].set_index("Status"), color="#c8a96e")
        with ch2:
            st.subheader(t("del_status_chart"))
            df_dc = pd.DataFrame({"Status":list(del_counts.keys()),"Count":list(del_counts.values())})
            st.bar_chart(df_dc[df_dc["Count"]>0].set_index("Status"), color="#8B5E3C")
        st.divider()
        pend = [r for r in orders if r[0] and str(r[13]).strip() in ("Pending","Ready")]
        if pend:
            st.subheader(t("pending_alert"))
            df_pend = pd.DataFrame(pend)[[0,2,3,5,8,9,10,12,13]]
            df_pend.columns = ["Order ID","Customer","Phone","Item","Delivery Date","Total","Paid","Pay Status","Delivery"]
            df_pend["Total"]         = df_pend["Total"].apply(lambda x: f"₹{safe_num(x):,.0f}")
            df_pend["Paid"]          = df_pend["Paid"].apply(lambda x: f"₹{safe_num(x):,.0f}")
            df_pend["Delivery Date"] = df_pend["Delivery Date"].apply(fmt_date)
            st.dataframe(df_pend, use_container_width=True, hide_index=True)
        else:
            st.success(t("no_pending"))
    except Exception as e:
        st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# ORDERS
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Orders":
    st.header(t("new_order"))
    # ── Customer Info ──
    st.subheader("👤 Customer Details")
    col1,col2 = st.columns(2)
    with col1:
        st.text_input(t("cust_name"),  key="customer_name")
        st.text_input(t("phone"),      key="phone_number")
        st.text_area(t("address"),     key="address", height=70)
    with col2:
        st.date_input(t("order_date"),    key="order_date",    value=date.today())
        st.date_input(t("delivery_date"), key="delivery_date", value=date.today())
        st.selectbox(t("del_status"),     ["Pending","Ready","Delivered"], key="delivery_status")

    # ── Jewellery Details ──
    st.subheader(t("jewel_details"))
    col3,col4 = st.columns(2)
    with col3:
        st.text_input(t("item_name"),    key="item_name")
        st.text_area(t("item_desc"),     key="item_desc",   height=70)
        st.selectbox(t("metal_type"),    ["Gold","Silver","Platinum","Rose Gold","Other"], key="metal_type")
        st.selectbox(t("purity"),        ["24K","22K","18K","14K","925 Silver","950 Platinum","Other"], key="purity")
    with col4:
        st.number_input(t("weight"),         key="weight",         min_value=0.0, step=0.1, format="%.2f")
        st.number_input(t("making_charges"), key="making_charges", min_value=0,   step=100)
        st.number_input(t("cost_price"),     key="cost_price",     min_value=0,   step=100)
        st.text_input(t("stone_details"),    key="stone_details")

    # ── Payment Info ──
    st.subheader("💰 Payment Details")
    col5,col6 = st.columns(2)
    with col5:
        st.number_input(t("total_amt"), key="total_amount", min_value=0, step=100)
    with col6:
        st.number_input(t("advance"),   key="advance_paid", min_value=0, step=100)

    total = st.session_state.get("total_amount",0)
    adv   = st.session_state.get("advance_paid",0)
    mc    = st.session_state.get("making_charges",0)
    cp    = st.session_state.get("cost_price",0)
    profit= total - cp - mc if total>0 else 0
    col_a, col_b = st.columns(2)
    col_a.markdown(f"### {t('balance_preview')}: ₹{max(total-adv,0):,}")
    if cp > 0:
        col_b.markdown(f"### 📈 Est. Profit: ₹{profit:,}")

    if st.button(t("save_order"), type="primary"):
        cname  = st.session_state["customer_name"].strip()
        phone  = st.session_state["phone_number"].strip()
        addr   = st.session_state["address"].strip()
        iname  = st.session_state["item_name"].strip()
        idesc  = st.session_state["item_desc"].strip()
        mtype  = st.session_state["metal_type"]
        purity = st.session_state["purity"]
        weight = st.session_state["weight"]
        making = st.session_state["making_charges"]
        cost   = st.session_state["cost_price"]
        stone  = st.session_state["stone_details"].strip()
        total_a= st.session_state["total_amount"]
        adv_a  = st.session_state["advance_paid"]
        odate  = st.session_state["order_date"]
        ddate  = st.session_state["delivery_date"]
        dstatus= st.session_state["delivery_status"]
        bal_a  = total_a - adv_a

        if not cname:       st.error(f"{t('cust_name')} required.")
        elif not phone:     st.error(f"{t('phone')} required.")
        elif not iname:     st.error(f"{t('item_name')} required.")
        elif adv_a>total_a: st.error("Advance > Total!")
        else:
            try:
                wb=load_workbook(EXCEL_FILE); ws_o=wb["Orders"]; ws_p=wb["Payments"]
                last_id=ws_o.cell(ws_o.max_row,1).value
                last_n =int(str(last_id).replace("O","")) if last_id and str(last_id).startswith("O") else 0
                oid=f"O{last_n+1:03d}"; cid=f"C{last_n+1:03d}"; nr=ws_o.max_row+1
                pst="Paid" if bal_a<=0 else ("Partially Paid" if adv_a>0 else "Pending")
                # Columns 1-15 original + 16-21 new jewellery fields
                for c,v in zip(range(1,16),[oid,cid,cname,phone,addr,iname,idesc,
                                             str(odate),str(ddate),total_a,adv_a,bal_a,pst,dstatus,""]):
                    ws_o.cell(nr,c).value=v
                # Extra jewellery columns 16-21
                for c,v in zip(range(16,22),[mtype,purity,weight,making,cost,stone]):
                    ws_o.cell(nr,c).value=v
                if adv_a>0:
                    lp=ws_p.cell(ws_p.max_row,1).value
                    lpn=int(str(lp).replace("P","")) if lp and str(lp).startswith("P") else 0
                    pr=ws_p.max_row+1
                    for c,v in zip(range(1,9),[f"P{lpn+1:03d}",oid,str(odate),adv_a,"Cash","",cname,iname]):
                        ws_p.cell(pr,c).value=v
                wb.save(EXCEL_FILE)
                profit_msg = f" | Est. Profit: ₹{profit:,}" if cost>0 else ""
                st.success(f"✅ {oid} saved! Advance: ₹{adv_a:,} | Balance: ₹{bal_a:,}{profit_msg}")
            except Exception as e: st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# PAYMENTS
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Payments":
    st.header(t("record_payment"))
    shop=get_shop_info()
    try:
        wb=load_workbook(EXCEL_FILE,read_only=True,data_only=True)
        pending=[]
        for row in wb["Orders"].iter_rows(min_row=2,values_only=True):
            if not row[0]: continue
            ta=safe_num(row[9]); pa=safe_num(row[10]); b=ta-pa
            if b>0:
                pending.append({"Order ID":row[0],"Customer":row[2],"Phone":row[3],"Item":row[5],
                                 "Total":f"₹{ta:,.0f}","Paid":f"₹{pa:,.0f}","Balance Due":f"₹{b:,.0f}","Status":row[12]})
        wb.close()
        if pending:
            st.subheader(t("orders_balance"))
            st.dataframe(pd.DataFrame(pending),use_container_width=True,hide_index=True)
        else: st.success(t("no_balance"))
    except Exception as e: st.warning(f"{e}")

    st.divider(); st.subheader(t("enter_payment"))
    c1,c2=st.columns(2)
    with c1:
        st.text_input(t("order_id"),   key="pay_order_id")
        st.number_input(t("amount_now"),key="amount_paid",min_value=0,step=100)
        st.date_input(t("payment_date"),key="payment_date",value=date.today())
    with c2:
        st.selectbox(t("payment_mode"),["Cash","UPI","GPay","PhonePe","Bank Transfer"],key="payment_mode")
        st.text_input(t("txn_id"),     key="transaction_id")

    if st.button(t("save_payment"),type="primary"):
        pay_oid=st.session_state["pay_order_id"].strip()
        amt    =st.session_state["amount_paid"]
        mode   =st.session_state["payment_mode"]
        txn    =st.session_state["transaction_id"].strip()
        pdate  =st.session_state["payment_date"]
        if not pay_oid: st.error("Order ID required.")
        elif amt<=0:    st.error("Amount must be > 0.")
        else:
            try:
                wb=load_workbook(EXCEL_FILE); ws_p=wb["Payments"]; ws_o=wb["Orders"]
                found=False; cname=iname=phone_no=ddate=""
                for row in ws_o.iter_rows(min_row=2):
                    if str(row[0].value)==pay_oid:
                        found=True; cname=row[2].value; iname=row[5].value
                        phone_no=row[3].value; ddate=row[8].value
                        ta=safe_num(row[9].value); cp=safe_num(row[10].value); cb=ta-cp
                        if amt>cb: st.error(f"₹{amt:,} exceeds balance ₹{cb:,.0f}."); wb.close(); st.stop()
                        np2=cp+amt; nb=ta-np2; ns="Paid" if nb<=0 else "Partially Paid"
                        row[10].value=np2; row[11].value=nb; row[12].value=ns; break
                if not found: st.error(f"'{pay_oid}' not found."); wb.close(); st.stop()
                lp=ws_p.cell(ws_p.max_row,1).value
                lpn=int(str(lp).replace("P","")) if lp and str(lp).startswith("P") else 0
                pr=ws_p.max_row+1
                for c,v in zip(range(1,9),[f"P{lpn+1:03d}",pay_oid,str(pdate),amt,mode,txn,cname,iname]):
                    ws_p.cell(pr,c).value=v
                wb.save(EXCEL_FILE)
                if nb<=0: st.success(f"✅ Fully PAID! 🎉 {pay_oid}")
                else:      st.success(f"✅ Paid ₹{np2:,.0f} | Remaining ₹{nb:,.0f}")
                if nb>0:   whatsapp_widget(pay_oid,cname,iname,nb,ddate,shop["name"],key_suffix="pay")
            except Exception as e: st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# EDIT / DELETE
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Edit / Delete Order":
    st.header(t("edit_delete"))
    st.info(t("edit_info"))
    eid_in=st.text_input(t("search_order"))
    if eid_in:
        eid=eid_in.strip().upper()
        try:
            wb=load_workbook(EXCEL_FILE,data_only=True); ws_o=wb["Orders"]
            trow=None; tidx=None
            for idx,row in enumerate(ws_o.iter_rows(min_row=2),start=2):
                if str(row[0].value).upper()==eid: trow=row; tidx=idx; break
            wb.close()
            if not trow: st.warning(f"'{eid}' not found.")
            else:
                r=[c.value for c in trow]
                st.subheader(f"{t('editing')}: {eid}")
                tab1,tab2=st.tabs([f"✏️ {'Edit' if st.session_state['lang']=='en' else 'திருத்து'}",
                                    f"🗑️ {'Delete' if st.session_state['lang']=='en' else 'நீக்கு'}"])
                with tab1:
                    c1,c2=st.columns(2)
                    with c1:
                        nc=st.text_input(t("cust_name"),  value=str(r[2] or ""),key="e_cn")
                        np=st.text_input(t("phone"),      value=str(r[3] or ""),key="e_ph")
                        na=st.text_area(t("address"),     value=str(r[4] or ""),key="e_ad",height=80)
                        ni=st.text_input(t("item_name"),  value=str(r[5] or ""),key="e_in")
                        nd=st.text_area(t("item_desc"),   value=str(r[6] or ""),key="e_id",height=80)
                    with c2:
                        nt=st.number_input(t("total_amt"),value=safe_num(r[9]), min_value=0,step=100,key="e_ta")
                        npd=st.number_input(t("advance"),  value=safe_num(r[10]),min_value=0,step=100,key="e_pd")
                        nds=st.selectbox(t("del_status"),["Pending","Ready","Delivered"],
                                          index=["Pending","Ready","Delivered"].index(str(r[13]).strip())
                                          if r[13] and str(r[13]).strip() in ["Pending","Ready","Delivered"] else 0,key="e_ds")
                    if st.button(t("save_changes"),type="primary"):
                        nb2=nt-npd; nps="Paid" if nb2<=0 else ("Partially Paid" if npd>0 else "Pending")
                        try:
                            wb2=load_workbook(EXCEL_FILE); ws2=wb2["Orders"]
                            for col,val in zip([3,4,5,6,7,10,11,12,13,14],[nc,np,na,ni,nd,nt,npd,nb2,nps,nds]):
                                ws2.cell(tidx,col).value=val
                            wb2.save(EXCEL_FILE); st.success(f"✅ {eid} updated!")
                        except Exception as e: st.error(f"❌ {e}")
                with tab2:
                    st.warning(t("delete_warn"))
                    conf=st.text_input(t("confirm_del"),key="del_conf")
                    if st.button(t("delete_order"),type="primary"):
                        if conf.strip().upper()!=eid: st.error("ID mismatch. Cancelled.")
                        else:
                            try:
                                wb3=load_workbook(EXCEL_FILE); ws3o=wb3["Orders"]; ws3p=wb3["Payments"]
                                ws3o.delete_rows(tidx)
                                to_del=[i for i,row in enumerate(ws3p.iter_rows(min_row=2,values_only=True),start=2)
                                        if row[1] and str(row[1]).upper()==eid]
                                for i in reversed(to_del): ws3p.delete_rows(i)
                                wb3.save(EXCEL_FILE); st.success(f"✅ {eid} deleted.")
                            except Exception as e: st.error(f"❌ {e}")
        except Exception as e: st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# CUSTOMER LEDGER
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Customer Ledger":
    st.header(t("cust_ledger"))
    st.info(t("ledger_info"))
    shop=get_shop_info()
    try:
        wb=load_workbook(EXCEL_FILE,read_only=True,data_only=True)
        ord_raw=list(wb["Orders"].iter_rows(min_row=2,values_only=True))
        pay_raw=list(wb["Payments"].iter_rows(min_row=2,values_only=True))
        wb.close()
        srch=st.text_input(t("search_cust"))
        if srch:
            s=srch.strip().lower()
            matched=[r for r in ord_raw if r[0] and(s in str(r[2]).lower() or s in str(r[3]).lower() or s in str(r[0]).lower())]
            if not matched: st.warning(t("no_cust"))
            else:
                for order in matched:
                    oid=order[0]; cname=order[2]; phone=order[3]; iname=order[5]
                    total=safe_num(order[9]); paid=safe_num(order[10]); bal=total-paid
                    pstat=order[12]; dstat=order[13]; ddate=order[8]
                    icon="🟢" if bal<=0 else("🟡" if paid>0 else"🔴")
                    with st.expander(f"{icon} {oid} — {cname} | {iname} | ₹{bal:,.0f}",expanded=True):
                        c1,c2,c3=st.columns(3)
                        c1.metric(t("total_amt"),  f"₹{total:,.0f}")
                        c2.metric(t("advance"),    f"₹{paid:,.0f}")
                        c3.metric(t("total_balance"),f"₹{bal:,.0f}")
                        st.markdown(f"""| | |\n|---|---|\n| **{t('phone')}** | {phone} |\n| **{t('item_name')}** | {iname} |\n| **{t('order_date')}** | {fmt_date(order[7])} |\n| **{t('delivery_date')}** | {fmt_date(ddate)} |\n| **Pay Status** | {pstat} |\n| **Delivery** | {dstat} |""")
                        cust_pay=[p for p in pay_raw if p[1] and str(p[1])==str(oid)]
                        if cust_pay:
                            st.markdown(f"**{t('pay_history')}:**")
                            n=len(cust_pay[0])
                            base=["Payment_ID","Order_ID","Date","Amount","Mode","Txn_ID","Customer","Item"]
                            cols=base[:n]+[f"Col{i}" for i in range(n-len(base))] if n>len(base) else base[:n]
                            df_p=pd.DataFrame(cust_pay,columns=cols)
                            show=[c for c in["Payment_ID","Date","Amount","Mode","Txn_ID"] if c in df_p.columns]
                            df_p=df_p[show]
                            if "Amount" in df_p.columns: df_p["Amount"]=df_p["Amount"].apply(lambda x:f"₹{safe_num(x):,.0f}")
                            if "Date"   in df_p.columns: df_p["Date"]  =df_p["Date"].apply(fmt_date)
                            st.dataframe(df_p,use_container_width=True,hide_index=True)
                        else: st.info(t("no_pay_rec"))
                        bc1,bc2=st.columns(2)
                        with bc1:
                            html_r=generate_pdf_receipt(order,cust_pay)
                            st.download_button(t("download_receipt"),data=html_r.encode("utf-8"),
                                               file_name=f"Receipt_{oid}.html",mime="text/html",key=f"pdf_{oid}")
                        with bc2:
                            if bal>0: whatsapp_widget(oid,cname,iname,bal,ddate,shop["name"],key_suffix=oid)
                        if bal>0: st.warning(f"{t('still_pending')} {cname}: ₹{bal:,.0f}")
                        else:     st.success(t("fully_paid"))
    except Exception as e: st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Statement":
    st.header(t("statement"))
    shop=get_shop_info()
    try:
        wb=load_workbook(EXCEL_FILE,read_only=True,data_only=True)
        od=list(wb["Orders"].iter_rows(values_only=True))
        pd2=list(wb["Payments"].iter_rows(values_only=True))
        wb.close()
        sid_in=st.text_input(t("search_order"))
        if sid_in:
            sid=sid_in.strip().upper()
            mo=[r for r in od[1:] if r[0] and str(r[0]).upper()==sid]
            mp=[r for r in pd2[1:] if r[1] and str(r[1]).upper()==sid]
            if not mo: st.warning(f"No order: {sid}")
            else:
                st.subheader(f"{t('order_details')} — {sid}")
                st.dataframe(pd.DataFrame(mo,columns=od[0]),use_container_width=True,hide_index=True)
                st.subheader(f"{t('payments_for')} {sid}")
                if mp:
                    base=["Payment_ID","Order_ID","Payment_Date","Amount_Paid","Payment_Mode","Transaction_ID","Customer_Name","Item_Name"]
                    n=len(mp[0]); ph=base[:n] if n<=len(base) else base+[f"Extra_{i}" for i in range(n-len(base))]
                    st.dataframe(pd.DataFrame(mp,columns=ph),use_container_width=True,hide_index=True)
                    ta=safe_num(mo[0][9]); pa=sum(safe_num(r[3]) for r in mp); ba=ta-pa
                    st.divider()
                    c1,c2,c3=st.columns(3)
                    c1.metric(t("total_amt"),f"₹{ta:,.0f}"); c2.metric(t("advance"),f"₹{pa:,.0f}"); c3.metric(t("total_balance"),f"₹{ba:,.0f}")
                    st.success(t("fully_paid")) if ba<=0 else st.warning(f"⚠️ ₹{ba:,.0f} pending.")
                else: st.info(t("no_pay_found"))
                html_r=generate_pdf_receipt(mo[0],mp)
                st.download_button(t("download_receipt"),data=html_r.encode("utf-8"),
                                   file_name=f"Receipt_{sid}.html",mime="text/html")
        else:
            st.info(f"👆 {t('search_order')}")
            st.divider(); st.subheader(t("quick_summary"))
            if len(od)>1:
                st.metric(t("total_in_system"),len(od)-1)
                st.markdown(f"**{t('recent_orders')}:**")
                st.dataframe(pd.DataFrame(od[1:6],columns=od[0]),use_container_width=True,hide_index=True)
    except Exception as e: st.error(f"❌ {e}")


# ══════════════════════════════════════════════════════════════════════════════
# ALERTS  —  Pending orders + Low balance WhatsApp
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Alerts":
    st.header("🔔 Alerts & Reminders")
    shop = get_shop_info()
    try:
        wb       = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        orders   = list(wb["Orders"].iter_rows(min_row=2, values_only=True))
        wb.close()

        today = date.today()

        # ── Overdue deliveries ──────────────────────────────────────────────
        overdue, due_today, pending_pay = [], [], []
        for r in orders:
            if not r[0]: continue
            bal      = safe_num(r[9]) - safe_num(r[10])
            ddate_str= fmt_date(r[8])
            dstatus  = str(r[13]).strip() if r[13] else "Pending"
            try:
                from datetime import datetime
                ddate_obj = datetime.strptime(ddate_str, "%Y-%m-%d").date() if ddate_str else None
            except: ddate_obj = None

            if dstatus != "Delivered":
                if ddate_obj and ddate_obj < today:
                    overdue.append(r)
                elif ddate_obj and ddate_obj == today:
                    due_today.append(r)

            if bal > 0 and dstatus != "Delivered":
                pending_pay.append(r)

        # ── Overdue ────────────────────────────────────────────────────────
        st.subheader(f"🚨 Overdue Deliveries ({len(overdue)})")
        if overdue:
            df_ov = pd.DataFrame(overdue)[[0,2,3,5,8,9,10,13]]
            df_ov.columns=["Order ID","Customer","Phone","Item","Delivery Date","Total","Paid","Delivery"]
            df_ov["Total"]=df_ov["Total"].apply(lambda x:f"₹{safe_num(x):,.0f}")
            df_ov["Paid"] =df_ov["Paid"].apply(lambda x:f"₹{safe_num(x):,.0f}")
            df_ov["Delivery Date"]=df_ov["Delivery Date"].apply(fmt_date)
            st.dataframe(df_ov, use_container_width=True, hide_index=True)
        else:
            st.success("✅ No overdue deliveries!")

        st.divider()

        # ── Due Today ──────────────────────────────────────────────────────
        st.subheader(f"📅 Deliveries Due Today ({len(due_today)})")
        if due_today:
            df_dt = pd.DataFrame(due_today)[[0,2,3,5,8,9,10,13]]
            df_dt.columns=["Order ID","Customer","Phone","Item","Delivery Date","Total","Paid","Delivery"]
            df_dt["Total"]=df_dt["Total"].apply(lambda x:f"₹{safe_num(x):,.0f}")
            df_dt["Paid"] =df_dt["Paid"].apply(lambda x:f"₹{safe_num(x):,.0f}")
            df_dt["Delivery Date"]=df_dt["Delivery Date"].apply(fmt_date)
            st.dataframe(df_dt, use_container_width=True, hide_index=True)
        else:
            st.success("✅ No deliveries due today!")

        st.divider()

        # ── Pending Payments with WhatsApp ─────────────────────────────────
        st.subheader(f"💰 Pending Balance Orders ({len(pending_pay)})")
        if pending_pay:
            st.info("Click 📲 next to each order to send a WhatsApp reminder.")
            for r in pending_pay:
                oid   = r[0]; cname = r[2]; phone = r[3]
                iname = r[5]; ddate = r[8]
                total = safe_num(r[9]); paid = safe_num(r[10]); bal = total - paid
                dstatus = str(r[13]).strip() if r[13] else "Pending"
                col1, col2, col3, col4, col5 = st.columns([1.5,2,1.5,1.5,2])
                col1.write(f"**{oid}**")
                col2.write(f"{cname}")
                col3.write(f"₹{bal:,.0f} due")
                col4.write(f"🚚 {dstatus}")
                with col5:
                    with st.popover("📲 Send Reminder"):
                        wa_num = st.text_input("WhatsApp No.", placeholder="9876543210",
                                               key=f"alert_wa_{oid}")
                        if wa_num:
                            clean = "".join(filter(str.isdigit, wa_num))
                            if len(clean) not in (10,12):
                                st.warning("Enter valid 10-digit number.")
                            else:
                                wa_url = whatsapp_link(wa_num, oid, cname, iname, bal, ddate, shop["name"])
                                st.link_button("✅ Open WhatsApp", wa_url)
        else:
            st.success("🎉 No pending balance orders!")

        st.divider()

        # ── Bulk summary ───────────────────────────────────────────────────
        st.subheader("📊 Alert Summary")
        c1,c2,c3 = st.columns(3)
        c1.metric("🚨 Overdue",       len(overdue))
        c2.metric("📅 Due Today",     len(due_today))
        c3.metric("💰 Pending Payment", len(pending_pay))

    except FileNotFoundError:
        st.error(f"❌ '{EXCEL_FILE}' not found.")
    except Exception as e:
        st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS + BACKUP & RESTORE
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Settings":
    st.header(t("settings"))
    try:
        wb=load_workbook(EXCEL_FILE,read_only=True,data_only=True)
        ws=wb["Settings"]
        row=list(ws.iter_rows(min_row=2,values_only=True))[0]
        wb.close()
        st.text_input(t("shop_name"),   value=str(row[0] or ""),disabled=True)
        st.text_input(t("shop_phone"),  value=str(row[1] or ""),disabled=True)
        st.text_input(t("shop_address"),value=str(row[2] or ""),disabled=True)
        st.info(t("settings_info"))
    except Exception: st.info(t("settings_info"))

    st.divider()
    st.subheader(t("backup_restore"))

    # ── Backup ──
    col1,col2=st.columns(2)
    with col1:
        st.markdown("#### 📥 Backup")
        if st.button(t("backup_now"),type="primary"):
            try:
                shutil.copy2(EXCEL_FILE, BACKUP_FILE)
                st.success(t("backup_success"))
            except Exception as e: st.error(f"❌ {e}")
        # Always offer download of backup if it exists
        if os.path.exists(BACKUP_FILE):
            with open(BACKUP_FILE,"rb") as f:
                st.download_button(t("backup_download"),data=f.read(),
                                   file_name=BACKUP_FILE,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # Also offer download of live file
        if os.path.exists(EXCEL_FILE):
            with open(EXCEL_FILE,"rb") as f:
                st.download_button("⬇️ Download Live Excel",data=f.read(),
                                   file_name=EXCEL_FILE,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_live")

    with col2:
        st.markdown(f"#### ♻️ {t('restore_title')}")
        st.warning(t("restore_warn"))
        uploaded=st.file_uploader(t("restore_upload"),type=["xlsx"],key="restore_upload")
        if uploaded and st.button(t("restore_btn"),type="primary"):
            try:
                with open(EXCEL_FILE,"wb") as f: f.write(uploaded.read())
                st.success(t("restore_success"))
                st.rerun()
            except Exception as e: st.error(f"❌ {e}")

# ══════════════════════════════════════════════════════════════════════════════
# HELP
# ══════════════════════════════════════════════════════════════════════════════
elif page_en_name == "Help":
    st.header(t("help"))
    st.markdown("""
### Features

| Feature | Where |
|---|---|
| 🔐 Login (admin/owner) | Login page |
| 🌐 Tamil / English switch | Sidebar |
| 📊 Dashboard charts | Dashboard |
| ⚠️ Pending delivery alerts | Dashboard |
| 💾 New order | Orders |
| 💳 Record payment | Payments |
| ✏️ Edit order | Edit / Delete Order |
| 🗑️ Delete order | Edit / Delete Order |
| 🔍 Search customer | Customer Ledger |
| 🖨️ Download Receipt (PDF) | Customer Ledger / Statement |
| 📲 WhatsApp Reminder | Customer Ledger |
| 📥 Backup Excel | Settings |
| ♻️ Restore from backup | Settings |

### Default Login
| Username | Password |
|---|---|
| admin | jewel123 |
| owner | shop456 |

> To change passwords, edit the `USERS` dict in app.py line ~190.

### Receipt as PDF
Download receipt → open in Chrome → Ctrl+P → Save as PDF
    """)